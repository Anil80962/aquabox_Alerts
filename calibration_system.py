#!/usr/bin/env python3
"""
Fluxgen Calibration Automation System v2.0
- Modbus RTU live meter reading (Reference ID:2, Calibration ID:1)
- Error %, Flow Deviation %, New K-Factor calculation
- PASS/FAIL validation
- Google Sheets export (new tab per serial number + master sheet)
- Local Excel export + master report
- Professional web dashboard
"""

import os, sys, json, time, struct, re, threading, glob
from datetime import datetime
from flask import Flask, request, jsonify, send_file
import minimalmodbus
import gspread
from google.oauth2.service_account import Credentials
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

app = Flask(__name__)

# ==================== PATHS ====================
BASE_DIR = os.path.dirname(os.path.abspath(__file__))
EXPORT_DIR = os.path.join(BASE_DIR, "exports")
HISTORY_FILE = os.path.join(BASE_DIR, "calibration_history.json")
MASTER_EXCEL = os.path.join(BASE_DIR, "Calibration_Master_Report.xlsx")
GOOGLE_CREDS = os.path.join(BASE_DIR, "credentials.json")
GOOGLE_SHEET_ID = "1bNP2WLn3GLjdZokq6g-MSa7EMJN3qnozSt5P-vuT5Xc"
LOGO_URL = "https://iili.io/qh9wrE7.png"

# ==================== TOLERANCES ====================
ERROR_TOL = 1.0      # +/- 1%
FLOW_DEV_TOL = 2.0   # +/- 2%

# ==================== MODBUS CONFIG ====================
MODBUS = {
    "baudrate": 9600, "parity": "N", "stopbits": 1, "timeout": 1.0,
    "ref_slave": 2, "cal_slave": 1,
    "start_reg": 0x1010, "num_regs": 12,
}

# ==================== STATE ====================
sessions = []
modbus_lock = threading.Lock()
live_reading = False
live_thread = None
live_data = {"ref_tot": 0.0, "cal_tot": 0.0, "ref_ok": False, "cal_ok": False, "last_read": "", "port": ""}

# ==================== HISTORY ====================
def load_history():
    global sessions
    if os.path.exists(HISTORY_FILE):
        try:
            with open(HISTORY_FILE) as f: sessions = json.load(f)
        except: sessions = []

def save_history():
    with open(HISTORY_FILE, "w") as f: json.dump(sessions, f, indent=2)

# ==================== CALCULATIONS ====================
def calculate(data):
    rv = float(data.get("ref_volume", 0))
    cv = float(data.get("cal_volume", 0))
    rf = float(data.get("ref_flow", 0))
    cf = float(data.get("cal_flow", 0))
    ek = float(data.get("existing_k", 1.0))

    err = ((cv - rv) / rv * 100) if rv else 0
    fdev = ((cf - rf) / rf * 100) if rf else 0
    nk = ek * (rv / cv) if cv else ek

    ep = abs(err) <= ERROR_TOL
    fp = abs(fdev) <= FLOW_DEV_TOL

    return {
        "ref_volume": rv, "cal_volume": cv, "ref_flow": rf, "cal_flow": cf,
        "existing_k": ek, "error_pct": round(err, 4), "flow_dev_pct": round(fdev, 4),
        "new_k": round(nk, 6), "error_pass": ep, "flow_pass": fp,
        "overall_pass": ep and fp, "status": "PASS" if (ep and fp) else "FAIL"
    }

# ==================== MODBUS ====================
def find_port():
    for p in ["/dev/ttyUSB0", "/dev/ttyUSB1", "/dev/ttyACM0", "/dev/serial0"]:
        if os.path.exists(p): return p
    return ""

def read_totalizer(slave_id):
    try:
        port = find_port()
        if not port: return 0.0, False
        with modbus_lock:
            inst = minimalmodbus.Instrument(port, slave_id)
            inst.serial.baudrate = MODBUS["baudrate"]
            inst.serial.parity = minimalmodbus.serial.PARITY_NONE
            inst.serial.stopbits = MODBUS["stopbits"]
            inst.serial.timeout = MODBUS["timeout"]
            inst.mode = minimalmodbus.MODE_RTU
            inst.clear_buffers_before_each_transaction = True
            regs = inst.read_registers(0x1010, 12, functioncode=4)
        if len(regs) < 12: return 0.0, False
        tot_int = struct.unpack(">i", struct.pack(">I", (regs[8]<<16)|regs[9]))[0]
        dec_flt = struct.unpack(">f", struct.pack(">I", (regs[10]<<16)|regs[11]))[0]
        return round(float(tot_int) + dec_flt, 4), True
    except Exception as e:
        print(f"[Modbus] Slave {slave_id} error: {e}")
        return 0.0, False

def read_both():
    r, rok = read_totalizer(MODBUS["ref_slave"])
    time.sleep(0.1)
    c, cok = read_totalizer(MODBUS["cal_slave"])
    live_data.update({"ref_tot": r, "cal_tot": c, "ref_ok": rok, "cal_ok": cok,
                      "last_read": datetime.now().strftime("%H:%M:%S"), "port": find_port()})
    return live_data.copy()

def live_loop():
    while live_reading:
        try: read_both()
        except: pass
        time.sleep(1)

def start_live():
    global live_reading, live_thread
    if live_reading: return False
    live_reading = True
    live_thread = threading.Thread(target=live_loop, daemon=True)
    live_thread.start()
    return True

def stop_live():
    global live_reading
    live_reading = False
    return True

# ==================== GOOGLE SHEETS EXPORT ====================
def export_gsheets(session):
    try:
        creds = Credentials.from_service_account_file(GOOGLE_CREDS, scopes=[
            "https://www.googleapis.com/auth/spreadsheets",
            "https://www.googleapis.com/auth/drive"
        ])
        gc = gspread.authorize(creds)
        sh = gc.open_by_key(GOOGLE_SHEET_ID)

        serial = session.get("serial_number", "UNKNOWN")
        ts = session.get("timestamp", datetime.now().strftime("%Y-%m-%d %H:%M:%S"))
        ek = session.get("existing_k", 1.0)
        mtype = session.get("meter_type", "")
        msize = session.get("meter_size", "")
        readings = session.get("readings", [])

        # ---- MASTER SHEET (Sheet1) ----
        try:
            mws = sh.worksheet("Sheet1")
        except:
            mws = sh.add_worksheet("Sheet1", 1000, 20)

        first = None
        try: first = mws.acell("A1").value
        except: pass

        if not first:
            mws.update("A1:P1", [["S.No","Date","Serial","Meter Type","Meter Size","Old K","Ref Vol","Test Vol","Error %",
                                   "Ref Flow","Test Flow","Flow Dev %","New K","Err Status","Flow Status","Overall"]])
            mws.format("A1:P1", {
                "backgroundColor": {"red":0.08,"green":0.4,"blue":0.75},
                "textFormat": {"bold":True,"foregroundColor":{"red":1,"green":1,"blue":1}},
                "horizontalAlignment": "CENTER"
            })

        vals = mws.get_all_values()
        nr = len(vals) + 1
        sno = nr - 1

        rows = []
        for rd in readings:
            r = rd.get("results", {})
            rows.append([sno, ts, serial, mtype, f"{msize} mm" if msize else "", ek,
                         r.get("ref_volume",0), r.get("cal_volume",0),
                         round(r.get("error_pct",0),4), r.get("ref_flow",0), r.get("cal_flow",0),
                         round(r.get("flow_dev_pct",0),4), round(r.get("new_k",0),6),
                         "PASS" if r.get("error_pass") else "FAIL",
                         "PASS" if r.get("flow_pass") else "FAIL",
                         r.get("status","FAIL")])
            sno += 1
        if rows:
            mws.update(f"A{nr}", rows)
            for i, row in enumerate(rows):
                rn = nr + i
                for ci in [14,15,16]:
                    cl = chr(64+ci)
                    v = row[ci-1]
                    bg = {"red":0.78,"green":0.9,"blue":0.79} if v=="PASS" else {"red":1,"green":0.8,"blue":0.82}
                    fg = {"red":0.18,"green":0.49,"blue":0.2} if v=="PASS" else {"red":0.78,"green":0.16,"blue":0.16}
                    mws.format(f"{cl}{rn}", {"backgroundColor":bg,"textFormat":{"bold":True,"foregroundColor":fg}})

        # ---- SUB-SHEET TAB ----
        tab = serial
        tabs = [w.title for w in sh.worksheets()]
        if tab in tabs:
            n = 2
            while f"{tab}_{n}" in tabs: n += 1
            tab = f"{tab}_{n}"

        ws = sh.add_worksheet(tab, 60, 12)

        # Logo centered (rows 1-4, cols B-H — bigger)
        ws.update_acell("B1", f'=IMAGE("{LOGO_URL}", 1)')
        ws.merge_cells("B1:H4")
        sh.batch_update({"requests":[
            {"updateDimensionProperties":{
                "range":{"sheetId":ws.id,"dimension":"ROWS","startIndex":0,"endIndex":4},
                "properties":{"pixelSize":35},"fields":"pixelSize"
            }}
        ]})
        ws.format("B1", {"horizontalAlignment": "CENTER", "verticalAlignment": "MIDDLE"})

        # Company name (row 5)
        ws.update("A5", [["FLUXGEN SUSTAINABLE TECHNOLOGIES"]])
        ws.merge_cells("A5:I5")
        ws.format("A5", {
            "textFormat":{"bold":True,"fontSize":14,"foregroundColor":{"red":0.05,"green":0.15,"blue":0.35}},
            "horizontalAlignment":"CENTER"
        })

        # Report title (row 6)
        ws.update("A6", [["Calibration Report"]])
        ws.merge_cells("A6:I6")
        ws.format("A6", {
            "textFormat":{"bold":True,"fontSize":11,"foregroundColor":{"red":0.08,"green":0.4,"blue":0.75}},
            "horizontalAlignment":"CENTER"
        })

        # Separator line (row 7 - thin)
        sh.batch_update({"requests":[{"updateDimensionProperties":{
            "range":{"sheetId":ws.id,"dimension":"ROWS","startIndex":6,"endIndex":7},
            "properties":{"pixelSize":6},"fields":"pixelSize"}}]})

        # Device Info (rows 8-12)
        ws.update("A8:B12", [
            ["Serial Number:", serial],
            ["Date:", ts],
            ["Meter Type:", mtype],
            ["Meter Size:", f"{msize} mm" if msize else ""],
            ["Existing K-Factor:", ek],
        ])
        ws.format("A8:A12", {"textFormat":{"bold":True,"fontSize":11,"foregroundColor":{"red":0.3,"green":0.3,"blue":0.3}}})
        ws.format("B8:B12", {"textFormat":{"bold":True,"fontSize":11}})

        # Readings Table
        hdr_row = 14
        ws.update(f"A{hdr_row}:I{hdr_row}", [["#","Ref Volume (L)","Test Volume (L)","Error %",
                   "Ref Flow (LPM)","Test Flow (LPM)","Flow Dev %","New K-Factor","Status"]])
        ws.format(f"A{hdr_row}:I{hdr_row}", {
            "backgroundColor":{"red":0.08,"green":0.4,"blue":0.75},
            "textFormat":{"bold":True,"foregroundColor":{"red":1,"green":1,"blue":1}},
            "horizontalAlignment":"CENTER"
        })

        data_rows = []
        for i, rd in enumerate(readings):
            r = rd.get("results", {})
            data_rows.append([i+1, r.get("ref_volume",0), r.get("cal_volume",0), round(r.get("error_pct",0),4),
                              r.get("ref_flow",0), r.get("cal_flow",0), round(r.get("flow_dev_pct",0),4),
                              round(r.get("new_k",0),6), r.get("status","FAIL")])

        ds = hdr_row + 1
        if data_rows:
            ws.update(f"A{ds}", data_rows)
            ws.format(f"A{ds}:I{ds+len(data_rows)-1}", {"horizontalAlignment":"CENTER"})
            for i, dr in enumerate(data_rows):
                rn = ds + i
                st = dr[-1]
                bg = {"red":0.78,"green":0.9,"blue":0.79} if st=="PASS" else {"red":1,"green":0.8,"blue":0.82}
                fg = {"red":0.18,"green":0.49,"blue":0.2} if st=="PASS" else {"red":0.78,"green":0.16,"blue":0.16}
                ws.format(f"I{rn}", {"backgroundColor":bg,"textFormat":{"bold":True,"fontSize":11,"foregroundColor":fg}})

        # ---- OVERALL SUMMARY ----
        if readings:
            errs = [rd["results"]["error_pct"] for rd in readings]
            fdevs = [rd["results"]["flow_dev_pct"] for rd in readings]
            nks = [rd["results"]["new_k"] for rd in readings]
            pc = sum(1 for rd in readings if rd["results"]["overall_pass"])
            fc = len(readings) - pc

            ae = sum(errs)/len(errs); me = max(abs(e) for e in errs)
            af = sum(fdevs)/len(fdevs); mf = max(abs(f) for f in fdevs)
            ak = sum(nks)/len(nks)
            kchange = ((ak - ek)/ek*100) if ek else 0
            opass = fc == 0
            ost = "PASS" if opass else "FAIL"

            sr = ds + len(data_rows) + 2
            ws.update(f"A{sr}:I{sr}", [["OVERALL CALIBRATION SUMMARY","","","","","","","",""]])
            ws.merge_cells(f"A{sr}:I{sr}")
            ws.format(f"A{sr}", {
                "backgroundColor":{"red":0.15,"green":0.15,"blue":0.3},
                "textFormat":{"bold":True,"fontSize":13,"foregroundColor":{"red":1,"green":1,"blue":1}},
                "horizontalAlignment":"CENTER"
            })

            r = sr + 1
            summary = [
                ["Parameter","Value","Tolerance","Status"],
                ["Avg Error %", f"{ae:.4f}%", f"+/- {ERROR_TOL}%", "PASS" if abs(ae)<=ERROR_TOL else "FAIL"],
                ["Max Error %", f"{me:.4f}%", f"+/- {ERROR_TOL}%", "PASS" if me<=ERROR_TOL else "FAIL"],
                ["Avg Flow Deviation %", f"{af:.4f}%", f"+/- {FLOW_DEV_TOL}%", "PASS" if abs(af)<=FLOW_DEV_TOL else "FAIL"],
                ["Max Flow Deviation %", f"{mf:.4f}%", f"+/- {FLOW_DEV_TOL}%", "PASS" if mf<=FLOW_DEV_TOL else "FAIL"],
                ["","","",""],
                ["Before Calibration (K-Factor)", str(ek), "", ""],
                ["After Calibration (K-Factor Avg)", f"{ak:.6f}", "", ""],
                ["K-Factor Change", f"{kchange:.4f}%", "", ""],
                ["Overall Deviation (Avg Error)", f"{ae:.4f}%", "", "PASS" if abs(ae)<=ERROR_TOL else "FAIL"],
                ["","","",""],
                ["Readings Passed", str(pc), f"out of {len(readings)}", ""],
                ["Readings Failed", str(fc), f"out of {len(readings)}", ""],
            ]
            ws.update(f"A{r}:D{r+len(summary)-1}", summary)
            ws.format(f"A{r}:D{r}", {
                "backgroundColor":{"red":0.88,"green":0.92,"blue":1.0},
                "textFormat":{"bold":True,"fontSize":10,"foregroundColor":{"red":0.08,"green":0.4,"blue":0.75}},
                "horizontalAlignment":"CENTER"
            })
            ws.format(f"A{r+1}:A{r+len(summary)-1}", {"textFormat":{"bold":True}})

            for i, s in enumerate(summary[1:], 1):
                if s[3] == "PASS":
                    ws.format(f"D{r+i}", {"backgroundColor":{"red":0.78,"green":0.9,"blue":0.79},
                        "textFormat":{"bold":True,"foregroundColor":{"red":0.18,"green":0.49,"blue":0.2}}})
                elif s[3] == "FAIL":
                    ws.format(f"D{r+i}", {"backgroundColor":{"red":1,"green":0.8,"blue":0.82},
                        "textFormat":{"bold":True,"foregroundColor":{"red":0.78,"green":0.16,"blue":0.16}}})

            # Big banner
            br = r + len(summary) + 1
            ws.update(f"A{br}:I{br}", [[f"OVERALL: {ost}","","","","","","","",""]])
            ws.merge_cells(f"A{br}:I{br}")
            bg = {"red":0.78,"green":0.9,"blue":0.79} if opass else {"red":1,"green":0.8,"blue":0.82}
            fg = {"red":0.18,"green":0.49,"blue":0.2} if opass else {"red":0.78,"green":0.16,"blue":0.16}
            ws.format(f"A{br}", {"backgroundColor":bg,"textFormat":{"bold":True,"fontSize":18,"foregroundColor":fg},"horizontalAlignment":"CENTER"})

            # Footer
            ws.update(f"A{br+2}", [["Fluxgen Sustainable Technologies - Calibration System v2.0"]])
            ws.merge_cells(f"A{br+2}:I{br+2}")
            ws.format(f"A{br+2}", {"textFormat":{"italic":True,"fontSize":9,"foregroundColor":{"red":0.6,"green":0.6,"blue":0.6}},"horizontalAlignment":"CENTER"})

        print(f"[GSheets] Exported '{serial}' -> tab '{tab}'")
        return True, tab
    except Exception as e:
        print(f"[GSheets] Error: {e}")
        return False, str(e)

# ==================== LOCAL EXCEL ====================
def generate_excel(session):
    os.makedirs(EXPORT_DIR, exist_ok=True)
    serial = session.get("serial_number", "UNKNOWN")
    fp = os.path.join(EXPORT_DIR, f"{serial}.xlsx")
    wb = Workbook(); ws = wb.active; ws.title = "Calibration Report"

    hf = Font(bold=True, color="FFFFFF", size=11)
    hfill = PatternFill(start_color="1565c0", end_color="1565c0", fill_type="solid")
    bdr = Border(left=Side("thin"), right=Side("thin"), top=Side("thin"), bottom=Side("thin"))
    ctr = Alignment(horizontal="center", vertical="center")

    for i, w in enumerate([8,20,16,14,14,14,12,14,14,12,14,12], 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    ws.merge_cells("A1:I1")
    ws["A1"].value = f"FLUXGEN - Calibration Report - {serial}"
    ws["A1"].font = Font(size=16, bold=True, color="1a237e"); ws["A1"].alignment = ctr
    mt = session.get("meter_type", "")
    ms = session.get("meter_size", "")
    ws["A2"].value = f"Date: {session.get('timestamp','')} | K-Factor: {session.get('existing_k','')} | Type: {mt} | Size: {ms} mm"

    r = 4
    headers = ["#","Ref Vol (L)","Test Vol (L)","Error %","Ref Flow","Test Flow","Flow Dev %","New K","Status"]
    for c, h in enumerate(headers, 1):
        cell = ws.cell(r, c, h); cell.font = hf; cell.fill = hfill; cell.alignment = ctr; cell.border = bdr

    for idx, rd in enumerate(session.get("readings", [])):
        res = rd.get("results", {})
        r += 1
        vals = [idx+1, res.get("ref_volume",0), res.get("cal_volume",0), round(res.get("error_pct",0),4),
                res.get("ref_flow",0), res.get("cal_flow",0), round(res.get("flow_dev_pct",0),4),
                round(res.get("new_k",0),6), res.get("status","FAIL")]
        for c, v in enumerate(vals, 1):
            cell = ws.cell(r, c, v); cell.alignment = ctr; cell.border = bdr
            if c == 9:
                cell.font = Font(bold=True, color="2e7d32" if v=="PASS" else "c62828")
                cell.fill = PatternFill(start_color="c8e6c9" if v=="PASS" else "ffcdd2", fill_type="solid")

    wb.save(fp)
    return fp, f"{serial}.xlsx"

def create_master_file():
    """Create an empty master Excel with headers."""
    wb = Workbook(); ws = wb.active; ws.title = "Records"
    headers = ["S.No","Date","Serial","Type","Size","Old K","Ref Vol","Test Vol","Error %",
               "Ref Flow","Test Flow","Flow Dev %","New K","Err","Flow","Overall"]
    hf = Font(bold=True, color="FFFFFF", size=11)
    hfill = PatternFill(start_color="1565c0", end_color="1565c0", fill_type="solid")
    ctr = Alignment(horizontal="center", vertical="center")
    bdr = Border(left=Side("thin"), right=Side("thin"), top=Side("thin"), bottom=Side("thin"))
    widths = [6,20,16,8,10,10,12,12,10,12,12,10,12,8,8,10]
    for c, (h, w) in enumerate(zip(headers, widths), 1):
        cell = ws.cell(1, c, h)
        cell.font = hf; cell.fill = hfill; cell.alignment = ctr; cell.border = bdr
        ws.column_dimensions[get_column_letter(c)].width = w
    ws.freeze_panes = "A2"
    wb.save(MASTER_EXCEL)
    print(f"[Excel] Created master file: {MASTER_EXCEL}")


def append_master(session):
    try:
        if os.path.exists(MASTER_EXCEL):
            wb = load_workbook(MASTER_EXCEL); ws = wb.active; nr = ws.max_row + 1; sno = nr - 1
        else:
            wb = Workbook(); ws = wb.active; ws.title = "Records"
            for c, h in enumerate(["S.No","Date","Serial","Type","Size","Old K","Ref Vol","Cal Vol","Error %",
                                    "Ref Flow","Cal Flow","Flow Dev %","New K","Err","Flow","Overall"], 1):
                ws.cell(1, c, h).font = Font(bold=True)
            nr = 2; sno = 1

        mt = session.get("meter_type", "")
        ms = session.get("meter_size", "")
        for rd in session.get("readings", []):
            r = rd.get("results", {})
            for c, v in enumerate([sno, session.get("timestamp",""), session.get("serial_number",""),
                                    mt, f"{ms} mm" if ms else "", session.get("existing_k",""),
                                    r.get("ref_volume",0), r.get("cal_volume",0),
                                    round(r.get("error_pct",0),4), r.get("ref_flow",0), r.get("cal_flow",0),
                                    round(r.get("flow_dev_pct",0),4), round(r.get("new_k",0),6),
                                    "PASS" if r.get("error_pass") else "FAIL",
                                    "PASS" if r.get("flow_pass") else "FAIL", r.get("status","FAIL")], 1):
                ws.cell(nr, c, v)
            nr += 1; sno += 1
        wb.save(MASTER_EXCEL)
    except Exception as e:
        print(f"[Excel] Master error: {e}")

# ==================== WEB UI ====================
HTML = r"""<!DOCTYPE html>
<html lang="en"><head>
<meta charset="UTF-8"><meta name="viewport" content="width=device-width,initial-scale=1">
<title>Fluxgen Calibration</title>
<style>
:root{--p:#1565c0;--s:#10b981;--d:#c62828;--w:#f57f17;--bg:#f0f2f5;--card:#fff;--txt:#1a1a2e;--mt:#6b7280;--bd:#e5e7eb}
*{box-sizing:border-box;margin:0;padding:0}
body{font-family:Inter,-apple-system,sans-serif;background:var(--bg);color:var(--txt);min-height:100vh}
.bar{background:linear-gradient(135deg,#0d47a1,#1565c0);padding:14px 20px;color:#fff;display:flex;align-items:center;justify-content:space-between}
.bar h1{font-size:17px;display:flex;align-items:center;gap:8px}.bar h1 svg{width:22px;height:22px;fill:#fff}
.bar .v{font-size:10px;opacity:.7}
.c{max-width:900px;margin:0 auto;padding:16px}
.card{background:var(--card);border-radius:12px;padding:22px;margin-bottom:14px;box-shadow:0 1px 3px rgba(0,0,0,.08);border:1px solid var(--bd)}
.ch{font-size:13px;font-weight:700;color:var(--p);text-transform:uppercase;letter-spacing:.5px;margin-bottom:16px;padding-bottom:10px;border-bottom:2px solid #e3f2fd;display:flex;align-items:center;gap:8px}
.ch svg{width:16px;height:16px;fill:var(--p)}
.fg{display:grid;grid-template-columns:1fr 1fr;gap:14px}
@media(max-width:600px){.fg{grid-template-columns:1fr}}
.fg label{display:block;font-size:11px;font-weight:600;color:var(--mt);text-transform:uppercase;letter-spacing:.4px;margin-bottom:5px}
.fg input,.fg select{width:100%;padding:10px 12px;border-radius:8px;border:1.5px solid var(--bd);font-size:14px;color:var(--txt);background:#fafafa;outline:none}
.fg input:focus{border-color:var(--p);box-shadow:0 0 0 3px rgba(21,101,192,.1)}
.btn{display:inline-flex;align-items:center;justify-content:center;gap:6px;padding:11px 20px;border:none;border-radius:10px;font-size:13px;font-weight:600;cursor:pointer;transition:.2s}
.btn:hover{transform:translateY(-1px);box-shadow:0 4px 12px rgba(0,0,0,.15)}
.btn:disabled{opacity:.4;cursor:not-allowed;transform:none}
.btn svg{width:14px;height:14px;fill:currentColor}
.bp{background:linear-gradient(135deg,var(--p),#1e88e5);color:#fff}
.bs{background:linear-gradient(135deg,var(--s),#43a047);color:#fff}
.bd{background:linear-gradient(135deg,var(--d),#e53935);color:#fff}
.bo{background:#fff;color:var(--p);border:1.5px solid var(--p)}
.br{display:flex;gap:8px;flex-wrap:wrap;margin-top:14px}
.rg{display:grid;grid-template-columns:repeat(auto-fit,minmax(160px,1fr));gap:10px;margin-bottom:14px}
.rc{padding:14px;border-radius:10px;text-align:center;border:1px solid var(--bd)}
.rc .l{font-size:10px;font-weight:600;color:var(--mt);text-transform:uppercase}
.rc .val{font-size:22px;font-weight:800;margin-top:3px}
.rc .u{font-size:11px;color:var(--mt);margin-top:2px}
.rc.pass{background:#e8f5e9;border-color:#a5d6a7}.rc.pass .val{color:#2e7d32}
.rc.fail{background:#ffebee;border-color:#ef9a9a}.rc.fail .val{color:#c62828}
.rc.n{background:#e3f2fd;border-color:#90caf9}.rc.n .val{color:var(--p)}
.sb{padding:18px;border-radius:12px;text-align:center;margin-bottom:14px}
.sb.pass{background:linear-gradient(135deg,#e8f5e9,#c8e6c9);border:2px solid #66bb6a}
.sb.fail{background:linear-gradient(135deg,#ffebee,#ffcdd2);border:2px solid #ef5350}
.sb .big{font-size:30px;font-weight:900}.sb.pass .big{color:#2e7d32}.sb.fail .big{color:#c62828}
.sb .inf{font-size:12px;color:var(--mt);margin-top:3px}
table{width:100%;border-collapse:collapse;font-size:12px;margin-top:10px}
th{background:#f0f4ff;padding:9px;text-align:center;font-weight:600;color:var(--p);border:1px solid var(--bd);font-size:10px;text-transform:uppercase}
td{padding:9px;text-align:center;border:1px solid var(--bd)}
.tp{background:#e8f5e9;color:#2e7d32;font-weight:700}.tf{background:#ffebee;color:#c62828;font-weight:700}
.hi{display:flex;align-items:center;justify-content:space-between;padding:10px 12px;border-radius:8px;border:1px solid var(--bd);margin-bottom:6px;cursor:pointer}
.hi:hover{background:#f8f9fa}
.hi .sn{font-weight:700;font-size:13px}.hi .dt{font-size:10px;color:var(--mt)}
.hb{font-size:10px;font-weight:700;padding:3px 8px;border-radius:5px}
.hb.p{background:#e8f5e9;color:#2e7d32}.hb.f{background:#ffebee;color:#c62828}
.hidden{display:none!important}
.toast{position:fixed;bottom:18px;left:50%;transform:translateX(-50%) translateY(100px);padding:10px 22px;border-radius:10px;font-size:12px;font-weight:500;color:#fff;z-index:100;max-width:90%;text-align:center;transition:transform .3s cubic-bezier(.175,.885,.32,1.275)}
.toast.show{transform:translateX(-50%) translateY(0)}
.spinner{display:inline-block;width:14px;height:14px;border:2px solid rgba(255,255,255,.2);border-top-color:currentColor;border-radius:50%;animation:sp .7s linear infinite;vertical-align:middle}
@keyframes sp{to{transform:rotate(360deg)}}
.mbx{display:grid;grid-template-columns:1fr 1fr;gap:12px;margin-bottom:12px}
.mbx>div{padding:14px;border-radius:10px;border:1.5px solid var(--bd)}
.mbx .tv{font-size:26px;font-weight:800;font-variant-numeric:tabular-nums}
.mbx .tl{font-size:10px;font-weight:700;color:var(--mt);text-transform:uppercase;margin-bottom:6px}
.mbx .ts{font-size:9px;padding:2px 6px;border-radius:4px;font-weight:600;display:inline-block;margin-top:4px}
.cap{background:#e3f2fd;border:1px solid #90caf9;border-radius:8px;padding:10px;margin-bottom:12px;display:flex;align-items:center;justify-content:space-between;flex-wrap:wrap;gap:6px}
.cap span{font-size:11px;color:var(--p);font-weight:600}
</style></head><body>
<div class="bar"><h1><img src="https://iili.io/qh9wrE7.png" style="height:30px;margin-right:10px;vertical-align:middle;border-radius:4px;">Fluxgen Calibration</h1><span class="v">v2.0</span></div>
<div class="c">

<!-- STEP 1 -->
<div class="card" id="s1">
<div class="ch"><svg viewBox="0 0 24 24"><path d="M12 2C6.48 2 2 6.48 2 12s4.48 10 10 10 10-4.48 10-10S17.52 2 12 2zm-2 15l-5-5 1.41-1.41L10 14.17l7.59-7.59L19 8l-9 9z"/></svg>Step 1 - Device Information</div>
<div class="fg">
<div><label>Serial Number</label><input id="sn" placeholder="e.g. FG24869L"></div>
<div><label>Existing K-Factor</label><input id="ek" type="number" step="any" value="1.0"></div>
<div><label>Meter Type</label><select id="mt"><option value="EMF">EMF - Electromagnetic Flow Meter</option><option value="USM">USM - Ultrasonic Flow Meter</option></select></div>
<div><label>Meter Size (mm)</label><input id="ms" type="text" placeholder="e.g. 25, 50, 100"></div>
</div>
<div class="br"><button class="btn bp" style="font-size:15px;padding:13px 30px" onclick="startCal()"><svg viewBox="0 0 24 24"><path d="M8 5v14l11-7z"/></svg>Start Calibration</button></div>
</div>

<!-- MODBUS PANEL -->
<div class="card hidden" id="mbPanel">
<div class="ch"><svg viewBox="0 0 24 24"><path d="M21 6h-2v9H6v2c0 .55.45 1 1 1h11l4 4V7c0-.55-.45-1-1-1z"/></svg>Modbus Live Readings<span style="margin-left:auto;font-size:10px;text-transform:none;letter-spacing:0">Port: <b id="mbPort">--</b> | <span id="mbSt" style="color:#f57f17">Stopped</span></span></div>
<div class="mbx">
<div style="background:#f8fffe"><div class="tl">Reference Meter (ID: <span id="rsi">2</span>)</div><div class="tv" style="color:var(--s)" id="lrt">0.0000</div><div style="font-size:10px;color:var(--mt)">Totalizer (L)</div><span class="ts" id="rSt" style="background:#e8f5e9;color:#2e7d32">WAIT</span></div>
<div style="background:#fff8f0"><div class="tl">Test Meter (ID: <span id="csi">1</span>)</div><div class="tv" style="color:var(--p)" id="lct">0.0000</div><div style="font-size:10px;color:var(--mt)">Totalizer (L)</div><span class="ts" id="cSt" style="background:#fff3e0;color:#e65100">WAIT</span></div>
</div>
<div style="font-size:10px;color:var(--mt);text-align:center;margin-bottom:8px">Last: <span id="lrt2">--</span></div>
<div class="br" style="margin-top:0"><button class="btn bp" id="lBtn" onclick="togLive()"><svg viewBox="0 0 24 24"><path d="M8 5v14l11-7z"/></svg>Start Live</button><button class="btn bo" onclick="oneRead()">Read Once</button></div>
</div>

<!-- STEP 2 -->
<div class="card hidden" id="s2">
<div class="ch"><svg viewBox="0 0 24 24"><path d="M3 17.25V21h3.75L17.81 9.94l-3.75-3.75L3 17.25z"/></svg>Step 2 - Readings<span style="margin-left:auto;font-size:11px;text-transform:none;letter-spacing:0">S/N: <b id="dSn">-</b> | K: <b id="dK">-</b> | <b id="dMt">-</b> <b id="dMs">-</b></span></div>
<div class="cap"><span>Auto-fill from Modbus</span><button class="btn bp" style="padding:7px 14px;font-size:11px;margin:0;width:auto" onclick="capMb()">Capture</button></div>
<div style="font-size:12px;font-weight:600;color:var(--mt);margin-bottom:8px">Reading #<span id="rn">1</span></div>
<div class="fg">
<div><label>Ref Volume (L)</label><input id="rv" type="number" step="any" placeholder="0.00"></div>
<div><label>Test Volume (L)</label><input id="cv" type="number" step="any" placeholder="0.00"></div>
<div><label>Ref Flow (LPM)</label><input id="rf" type="number" step="any" placeholder="0.00"></div>
<div><label>Test Flow (LPM)</label><input id="cf" type="number" step="any" placeholder="0.00"></div>
</div>
<div class="br">
<button class="btn bp" onclick="calc()"><svg viewBox="0 0 24 24"><path d="M9 16.17L4.83 12l-1.42 1.41L9 19 21 7l-1.41-1.41z"/></svg>Calculate</button>
<button class="btn bo" onclick="addRd()"><svg viewBox="0 0 24 24"><path d="M19 13h-6v6h-2v-6H5v-2h6V5h2v6h6v2z"/></svg>Add Reading</button>
<button class="btn bd" onclick="reset()">Reset</button>
</div></div>

<!-- STEP 3 -->
<div class="card hidden" id="s3">
<div class="ch"><svg viewBox="0 0 24 24"><path d="M19 3H5c-1.1 0-2 .9-2 2v14c0 1.1.9 2 2 2h14c1.1 0 2-.9 2-2V5c0-1.1-.9-2-2-2zM9 17H7v-7h2v7zm4 0h-2V7h2v10zm4 0h-2v-4h2v4z"/></svg>Results</div>
<div id="sBan" class="sb pass"><div class="big" id="oSt">PASS</div><div class="inf" id="sInf">All within tolerance</div></div>
<div class="rg" id="rGrid"></div>
<div id="rtWrap"></div>
<div class="br">
<button class="btn bs" style="font-size:13px;padding:11px 18px" onclick="doExport()"><svg viewBox="0 0 24 24"><path d="M19 9h-4V3H9v6H5l7 7 7-7zM5 18v2h14v-2H5z"/></svg>Export Excel</button>
<button class="btn bp" style="font-size:13px;padding:11px 18px" onclick="gsOnly()"><svg viewBox="0 0 24 24"><path d="M19 9h-4V3H9v6H5l7 7 7-7zM5 18v2h14v-2H5z"/></svg>Google Sheets</button>
<button class="btn bo" style="font-size:13px;padding:11px 18px" onclick="newCal()">New Calibration</button>
</div></div>

<!-- HISTORY -->
<div class="card">
<div class="ch"><svg viewBox="0 0 24 24"><path d="M13 3a9 9 0 00-9 9H1l3.89 3.89.07.14L9 12H6c0-3.87 3.13-7 7-7s7 3.13 7 7-3.13 7-7 7c-1.93 0-3.68-.79-4.94-2.06l-1.42 1.42A8.954 8.954 0 0013 21a9 9 0 000-18z"/></svg>History</div>
<div class="br" style="margin-top:0;margin-bottom:12px"><button class="btn bs" onclick="dlMaster()"><svg viewBox="0 0 24 24"><path d="M19 9h-4V3H9v6H5l7 7 7-7zM5 18v2h14v-2H5z"/></svg>Download Master Report</button></div>
<div id="hList"><div style="text-align:center;color:var(--mt);padding:14px;font-size:12px">No records yet</div></div>
</div>
</div>

<div class="toast" id="toast"></div>
<script>
var cur=null,rds=[],sRef=null,sCal=null,liveOn=false,liveInt=null;

function toast(m,bg){var t=document.getElementById('toast');t.textContent=m;t.style.background=bg||'#1565c0';t.className='toast show';setTimeout(function(){t.className='toast'},3500)}
function rc(l,v,u,p){var c='rc '+(p===true?'pass':p===false?'fail':'n');return'<div class="'+c+'"><div class="l">'+l+'</div><div class="val">'+v+'</div><div class="u">'+u+'</div></div>'}

function startCal(){
  var s=document.getElementById('sn').value.trim(),k=document.getElementById('ek').value;
  if(!s){toast('Enter Serial Number','#c62828');return}
  var mt=document.getElementById('mt').value;
  var ms=document.getElementById('ms').value.trim();
  cur={serial_number:s,existing_k:parseFloat(k),meter_type:mt,meter_size:ms};rds=[];sRef=null;sCal=null;
  document.getElementById('dSn').textContent=s;document.getElementById('dK').textContent=k;
  document.getElementById('dMt').textContent=mt;document.getElementById('dMs').textContent=ms?ms+'mm':'';
  document.getElementById('rn').textContent='1';
  ['s2','s3','mbPanel'].forEach(function(id){document.getElementById(id).classList.remove('hidden')});
  clr();chkMb();document.getElementById('mbPanel').scrollIntoView({behavior:'smooth'});
  toast('Calibration started for '+s)
}
function clr(){['rv','cv','rf','cf'].forEach(function(id){document.getElementById(id).value=''})}
function getSess(){return{serial_number:cur.serial_number,existing_k:cur.existing_k,meter_type:cur.meter_type,meter_size:cur.meter_size,readings:rds,timestamp:new Date().toLocaleString()}}

function calc(){
  var d={ref_volume:parseFloat(document.getElementById('rv').value)||0,cal_volume:parseFloat(document.getElementById('cv').value)||0,
         ref_flow:parseFloat(document.getElementById('rf').value)||0,cal_flow:parseFloat(document.getElementById('cf').value)||0,existing_k:cur.existing_k};
  if(!d.ref_volume){toast('Enter Ref Volume','#c62828');return}
  if(!d.cal_volume){toast('Enter Test Volume','#c62828');return}
  fetch('/api/calc',{method:'POST',headers:{'Content-Type':'application/json'},body:JSON.stringify(d)})
  .then(function(r){return r.json()}).then(function(x){
    if(x.success){rds.push({inputs:d,results:x.results});showRes(x.results);toast('Done: '+x.results.status,x.results.overall_pass?'#2e7d32':'#c62828')}
  })
}
function addRd(){
  var d={ref_volume:parseFloat(document.getElementById('rv').value)||0,cal_volume:parseFloat(document.getElementById('cv').value)||0,
         ref_flow:parseFloat(document.getElementById('rf').value)||0,cal_flow:parseFloat(document.getElementById('cf').value)||0,existing_k:cur.existing_k};
  if(!d.ref_volume||!d.cal_volume){toast('Fill current reading first','#c62828');return}
  fetch('/api/calc',{method:'POST',headers:{'Content-Type':'application/json'},body:JSON.stringify(d)})
  .then(function(r){return r.json()}).then(function(x){
    if(x.success){rds.push({inputs:d,results:x.results});clr();document.getElementById('rn').textContent=rds.length+1;toast('Reading #'+rds.length+' saved')}
  })
}
function showRes(lr){
  document.getElementById('s3').classList.remove('hidden');
  var b=document.getElementById('sBan');b.className='sb '+(lr.overall_pass?'pass':'fail');
  document.getElementById('oSt').textContent=lr.status;
  document.getElementById('sInf').textContent=lr.overall_pass?'All parameters within tolerance':'One or more out of tolerance';
  document.getElementById('rGrid').innerHTML=
    rc('Error %',lr.error_pct.toFixed(4)+'%','Tol: +/-1%',lr.error_pass)+
    rc('Flow Dev',lr.flow_dev_pct.toFixed(4)+'%','Tol: +/-2%',lr.flow_pass)+
    rc('New K',lr.new_k.toFixed(6),'Old: '+lr.existing_k,null)+
    rc('Ref Vol',lr.ref_volume+' L','Reference',null)+
    rc('Test Vol',lr.cal_volume+' L','Test Meter',null)+
    rc('Readings',rds.length,'Total',null);
  if(rds.length>0){
    var t='<table><thead><tr><th>#</th><th>Ref Vol</th><th>Test Vol</th><th>Error%</th><th>Ref Flow</th><th>Test Flow</th><th>Flow Dev%</th><th>New K</th><th>Status</th></tr></thead><tbody>';
    rds.forEach(function(r,i){var s=r.results;t+='<tr><td>'+(i+1)+'</td><td>'+s.ref_volume+'</td><td>'+s.cal_volume+'</td><td>'+s.error_pct.toFixed(4)+'</td><td>'+s.ref_flow+'</td><td>'+s.cal_flow+'</td><td>'+s.flow_dev_pct.toFixed(4)+'</td><td>'+s.new_k.toFixed(6)+'</td><td class="'+(s.overall_pass?'tp':'tf')+'">'+s.status+'</td></tr>'});
    t+='</tbody></table>';document.getElementById('rtWrap').innerHTML=t}
  document.getElementById('s3').scrollIntoView({behavior:'smooth'})
}
function doExport(){
  if(!cur||!rds.length){toast('No data','#c62828');return}
  toast('Exporting to Excel...','#1565c0');
  fetch('/api/export',{method:'POST',headers:{'Content-Type':'application/json'},body:JSON.stringify(getSess())})
  .then(function(r){if(r.ok)return r.blob().then(function(b){var u=URL.createObjectURL(b),a=document.createElement('a');a.href=u;a.download=cur.serial_number+'.xlsx';a.click();URL.revokeObjectURL(u);toast('Excel exported: '+cur.serial_number+'.xlsx','#2e7d32')});else toast('Export failed','#c62828')})
}
function gsOnly(){
  if(!cur||!rds.length){toast('No data','#c62828');return}
  toast('Exporting to Google Sheets...','#1565c0');
  fetch('/api/export_gs',{method:'POST',headers:{'Content-Type':'application/json'},body:JSON.stringify(getSess())})
  .then(function(r){return r.json()}).then(function(d){toast(d.message,d.success?'#2e7d32':'#c62828')})
}
function reset(){cur=null;rds=[];sRef=null;sCal=null;if(liveOn)togLive();['s2','s3','mbPanel'].forEach(function(id){document.getElementById(id).classList.add('hidden')});clr();document.getElementById('sn').value='';document.getElementById('ek').value='1.0';window.scrollTo({top:0,behavior:'smooth'})}
function newCal(){if(cur&&rds.length){fetch('/api/save',{method:'POST',headers:{'Content-Type':'application/json'},body:JSON.stringify(getSess())}).then(function(){loadH()})}reset()}
function dlMaster(){fetch('/api/master').then(function(r){if(r.ok)return r.blob().then(function(b){var u=URL.createObjectURL(b),a=document.createElement('a');a.href=u;a.download='Calibration_Master.xlsx';a.click();URL.revokeObjectURL(u);toast('Master report downloaded!','#2e7d32')});else toast('No data in master report yet. Export a calibration first.','#f57f17')}).catch(function(){toast('Download failed','#c62828')})}
function loadH(){fetch('/api/history').then(function(r){return r.json()}).then(function(d){var el=document.getElementById('hList');if(!d.s||!d.s.length){el.innerHTML='<div style="text-align:center;color:var(--mt);padding:14px;font-size:12px">No records</div>';return}
  var h='';d.s.slice().reverse().forEach(function(s){var lr=s.readings&&s.readings.length?s.readings[s.readings.length-1]:null;var st=lr?lr.results.status:'N/A';var p=st==='PASS';
  h+='<div class="hi"><div><div class="sn">'+s.serial_number+'</div><div class="dt">'+s.timestamp+' | '+(s.meter_type||'')+' '+(s.meter_size?s.meter_size+'mm':'')+' | K:'+s.existing_k+'</div></div><span class="hb '+(p?'p':'f')+'">'+st+'</span></div>'});el.innerHTML=h})}

// Modbus
function chkMb(){fetch('/api/mb/status').then(function(r){return r.json()}).then(function(d){document.getElementById('mbPort').textContent=d.port||'None';document.getElementById('rsi').textContent=d.cfg.ref;document.getElementById('csi').textContent=d.cfg.cal})}
function updMb(d){document.getElementById('lrt').textContent=d.ref_tot.toFixed(4);document.getElementById('lct').textContent=d.cal_tot.toFixed(4);document.getElementById('lrt2').textContent=d.last_read||'--';
  var rs=document.getElementById('rSt'),cs=document.getElementById('cSt');
  if(d.ref_ok){rs.textContent='OK';rs.style.background='#e8f5e9';rs.style.color='#2e7d32'}else{rs.textContent='ERR';rs.style.background='#ffebee';rs.style.color='#c62828'}
  if(d.cal_ok){cs.textContent='OK';cs.style.background='#e8f5e9';cs.style.color='#2e7d32'}else{cs.textContent='ERR';cs.style.background='#ffebee';cs.style.color='#c62828'}}
function togLive(){
  if(liveOn){fetch('/api/mb/stop',{method:'POST'});clearInterval(liveInt);liveOn=false;var b=document.getElementById('lBtn');b.innerHTML='<svg viewBox="0 0 24 24"><path d="M8 5v14l11-7z"/></svg>Start Live';b.className='btn bp';document.getElementById('mbSt').textContent='Stopped';document.getElementById('mbSt').style.color='#f57f17'}
  else{fetch('/api/mb/start',{method:'POST'}).then(function(r){return r.json()}).then(function(d){if(d.success){liveOn=true;var b=document.getElementById('lBtn');b.innerHTML='<svg viewBox="0 0 24 24"><path d="M6 6h12v12H6z"/></svg>Stop';b.className='btn bd';document.getElementById('mbSt').textContent='Running';document.getElementById('mbSt').style.color='#2e7d32';liveInt=setInterval(pollMb,1000)}})}
}
function pollMb(){fetch('/api/mb/status').then(function(r){return r.json()}).then(function(d){if(d.data)updMb(d.data)})}
function oneRead(){toast('Reading...','#1565c0');fetch('/api/mb/read',{method:'POST'}).then(function(r){return r.json()}).then(function(d){if(d.success){updMb(d.data);toast('Done','#2e7d32')}else toast('Failed: '+d.message,'#c62828')})}
function capMb(){toast('Capturing...','#1565c0');fetch('/api/mb/read',{method:'POST'}).then(function(r){return r.json()}).then(function(d){if(d.success&&d.data){if(sRef===null){sRef=d.data.ref_tot;sCal=d.data.cal_tot;toast('Start captured! Run water, then capture again.','#1565c0');return}
  var rv=Math.abs(d.data.ref_tot-sRef),cv=Math.abs(d.data.cal_tot-sCal);document.getElementById('rv').value=rv.toFixed(4);document.getElementById('cv').value=cv.toFixed(4);sRef=null;sCal=null;toast('Volumes: Ref '+rv.toFixed(4)+'L, Cal '+cv.toFixed(4)+'L','#2e7d32')}else toast('Failed','#c62828')})}

loadH();
</script></body></html>"""

# ==================== API ROUTES ====================
@app.route("/")
def index(): return HTML

@app.route("/api/calc", methods=["POST"])
def api_calc():
    d = request.get_json(); return jsonify({"success": True, "results": calculate(d)})

@app.route("/api/export", methods=["POST"])
def api_export():
    s = request.get_json()
    append_master(s)
    fp, fn = generate_excel(s)
    return send_file(fp, as_attachment=True, download_name=fn, mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")

@app.route("/api/export_gs", methods=["POST"])
def api_export_gs():
    s = request.get_json()
    ok, info = export_gsheets(s)
    return jsonify({"success": ok, "message": f"Exported to tab: {info}" if ok else f"Error: {info}"})

@app.route("/api/save", methods=["POST"])
def api_save():
    s = request.get_json(); sessions.append(s); save_history(); append_master(s); return jsonify({"success": True})

@app.route("/api/history")
def api_history(): return jsonify({"s": sessions})

@app.route("/api/master")
def api_master():
    if not os.path.exists(MASTER_EXCEL):
        create_master_file()
    if os.path.exists(MASTER_EXCEL):
        return send_file(MASTER_EXCEL, as_attachment=True, download_name="Calibration_Master.xlsx",
                         mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
    return jsonify({"success": False, "message": "No master report available"}), 404

@app.route("/api/mb/status")
def api_mb_status():
    return jsonify({"port":find_port(),"port_found":bool(find_port()),"live":live_reading,"data":live_data,
                    "cfg":{"ref":MODBUS["ref_slave"],"cal":MODBUS["cal_slave"],"baud":MODBUS["baudrate"]}})

@app.route("/api/mb/read", methods=["POST"])
def api_mb_read():
    try: return jsonify({"success": True, "data": read_both()})
    except Exception as e: return jsonify({"success": False, "message": str(e)})

@app.route("/api/mb/start", methods=["POST"])
def api_mb_start():
    return jsonify({"success": start_live(), "message": "Started" if start_live() else "Already running"})

@app.route("/api/mb/stop", methods=["POST"])
def api_mb_stop(): stop_live(); return jsonify({"success": True})

# Captive portal handlers
@app.route("/generate_204")
@app.route("/fwlink")
def captive(): return "", 204

# ==================== MAIN ====================
if __name__ == "__main__":
    sys.stdout.reconfigure(line_buffering=True)
    os.makedirs(EXPORT_DIR, exist_ok=True)
    load_history()
    if not os.path.exists(MASTER_EXCEL):
        create_master_file()
    print("=" * 50)
    print("  Fluxgen Calibration v2.0")
    print("  http://0.0.0.0:8085")
    print("=" * 50)
    app.run(host="0.0.0.0", port=8085, debug=False, use_reloader=False)
